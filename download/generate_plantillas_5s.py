"""
Generador de Plantillas 5S — Inventarios por cada S
=====================================================
Genera 4 plantillas Excel profesionales para los pasos 3 de cada S:
  - S1 Seiri:   Inventario de Innecesarios
  - S2 Seiton:  Inventario de Necesarios
  - S3 Seiso:   Inventario de Puntos de Suciedad
  - S4 Seiketsu: Inventario de Estándares Implantados
"""

import sys, os

XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from base import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

# Use professional palette for industrial/business context
use_palette_explicit("professional")

OUTPUT_DIR = "/home/z/my-project/download"

# ============================================================
# Color tokens per S (accent colors matching the 5S board)
# ============================================================
S_COLORS = {
    "S1": "8B5CF6",  # Purple - Seiri
    "S2": "EAB308",  # Yellow - Seiton
    "S3": "3B82F6",  # Blue - Seiso
    "S4": "F43F5E",  # Rose - Seiketsu
}


def _create_info_section(ws, s_code, s_title, last_col):
    """Create the project info header section at the top of each template."""
    # Title already at B2 from setup_sheet
    # Subtitle with S description
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=last_col)
    cell = ws.cell(row=3, column=2, value=f"{s_code} — {s_title}")
    cell.font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color=S_COLORS.get(s_code, PRIMARY))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 24

    # Info fields row
    info_labels = ["Empresa:", "Proyecto:", "Zona:", "Fecha:", "Responsable:"]
    info_row = 4
    col = 2
    for label in info_labels:
        cell_label = ws.cell(row=info_row, column=col, value=label)
        cell_label.font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
        cell_label.alignment = Alignment(horizontal="right", vertical="center")
        cell_value = ws.cell(row=info_row, column=col + 1, value="")
        cell_value.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600)
        cell_value.border = Border(bottom=Side(style="thin", color=NEUTRAL_200))
        col += 2

    ws.row_dimensions[4].height = 22
    # Spacer
    ws.row_dimensions[5].height = 8
    return 6  # Next available row (header row)


def _create_inventory_table(ws, header_row, headers, col_widths, data_validations=None, num_empty_rows=15):
    """Create the main inventory table with headers and empty data rows."""
    last_col = len(headers) + 1  # Starting from B=2

    # Style header row
    s_color = S_COLORS.get(ws.title[:2] if ws.title[:2] in S_COLORS else None, PRIMARY)

    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor=s_color)
        cell.font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=NEUTRAL_200))

    ws.row_dimensions[header_row].height = 32

    # Set column widths
    for col_idx, width in enumerate(col_widths, start=2):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write empty data rows
    data_start = header_row + 1
    for i in range(num_empty_rows):
        row_num = data_start + i
        # Row number
        ws.cell(row=row_num, column=2, value=i + 1)
        fill = fill_data_row(i)
        for col_idx in range(2, last_col + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = fill
            cell.font = font_body()
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if col_idx == 2:  # Nº column
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = 22

    # Apply data validations
    if data_validations:
        for dv in data_validations:
            ws.add_data_validation(dv)

    # Summary row
    summary_row = data_start + num_empty_rows
    ws.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=3)
    cell = ws.cell(row=summary_row, column=2, value="TOTAL ELEMENTOS")
    cell.font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=s_color)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    cell.border = Border(top=Side(style="medium", color=NEUTRAL_200))

    # Count formula
    count_col = 3  # Elemento column (C)
    formula_cell = ws.cell(row=summary_row, column=4)
    formula_cell.value = f'=COUNTA(C{data_start}:C{summary_row - 1})'
    formula_cell.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=s_color)
    formula_cell.alignment = Alignment(horizontal="center", vertical="center")
    formula_cell.fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    formula_cell.border = Border(top=Side(style="medium", color=NEUTRAL_200))

    # Fill remaining summary cells
    for col_idx in range(5, last_col + 1):
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
        cell.border = Border(top=Side(style="medium", color=NEUTRAL_200))

    ws.row_dimensions[summary_row].height = 26

    # Notes row
    notes_row = summary_row + 2
    ws.merge_cells(start_row=notes_row, start_column=2, end_row=notes_row, end_column=last_col)
    cell = ws.cell(row=notes_row, column=2, value="Notas: Completa esta plantilla durante el paso 3 del proceso 5S. Marca la clasificación y decisión para cada elemento identificado.")
    cell.font = font_caption()
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Legend row
    legend_row = notes_row + 1
    ws.merge_cells(start_row=legend_row, start_column=2, end_row=legend_row, end_column=last_col)
    cell = ws.cell(row=legend_row, column=2, value="Clasificación por colores de tarjeta: ROJA = Eliminar  AMARILLA = Reubicar/Revisar  VERDE = Mantener")
    cell.font = Font(name=FONT_NAME, size=9, bold=True, color=NEUTRAL_600)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    return summary_row


# ============================================================
# S1 — Inventario de Innecesarios (Seiri - Clasificar)
# ============================================================
def create_s1_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "S1 Innecesarios"

    headers = [
        "Nº", "Elemento / Objeto", "Ubicación", "Zona", "Cantidad",
        "Estado", "Frecuencia de Uso", "Clasificación",
        "Decisión", "Responsable", "Fecha Acción", "Observaciones"
    ]
    col_widths = [5, 22, 16, 12, 10, 12, 16, 16, 14, 14, 14, 24]

    last_col = len(headers) + 1
    setup_sheet(ws, title="Inventario de Innecesarios — SEIRI (Clasificar)", last_col=last_col)

    header_row = _create_info_section(ws, "S1", "Clasificar — Separar lo necesario de lo innecesario", last_col)

    # Data validations
    dv_estado = DataValidation(type="list", formula1='"Bueno,Regular,Malo"', allow_blank=True)
    dv_estado.prompt = "Selecciona el estado del elemento"
    dv_estado.promptTitle = "Estado"
    dv_estado.sqref = f"G{header_row + 1}:G{header_row + 15}"

    dv_frecuencia = DataValidation(type="list", formula1='"Diaria,Semanal,Mensual,Anual,Nunca"', allow_blank=True)
    dv_frecuencia.prompt = "Con qué frecuencia se usa"
    dv_frecuencia.promptTitle = "Frecuencia"
    dv_frecuencia.sqref = f"H{header_row + 1}:H{header_row + 15}"

    dv_clasificacion = DataValidation(type="list", formula1='"Innecesario,Dudoso"', allow_blank=True)
    dv_clasificacion.prompt = "Clasifica el elemento"
    dv_clasificacion.promptTitle = "Clasificación"
    dv_clasificacion.sqref = f"I{header_row + 1}:I{header_row + 15}"

    dv_decision = DataValidation(type="list", formula1='"Eliminar,Reubicar,Revisar"', allow_blank=True)
    dv_decision.prompt = "Qué hacer con el elemento"
    dv_decision.promptTitle = "Decisión"
    dv_decision.sqref = f"J{header_row + 1}:J{header_row + 15}"

    _create_inventory_table(ws, header_row, headers, col_widths, 
                           [dv_estado, dv_frecuencia, dv_clasificacion, dv_decision])

    # Freeze panes
    ws.freeze_panes = f"D{header_row + 1}"

    # Print setup
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

    wb.properties.creator = "Z.ai"
    path = os.path.join(OUTPUT_DIR, "S1_Inventario_Innecesarios_Seiri.xlsx")
    wb.save(path)
    print(f"✅ Generado: {path}")
    return path


# ============================================================
# S2 — Inventario de Necesarios (Seiton - Organizar)
# ============================================================
def create_s2_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "S2 Necesarios"

    headers = [
        "Nº", "Elemento / Objeto", "Ubicación Actual", "Zona", "Cantidad",
        "Frecuencia de Uso", "Ubicación Asignada", "Método Identificación",
        "Cercanía al Puesto", "Responsable", "Fecha Colocación", "Observaciones"
    ]
    col_widths = [5, 22, 16, 12, 10, 16, 16, 18, 16, 14, 14, 24]

    last_col = len(headers) + 1
    setup_sheet(ws, title="Inventario de Necesarios — SEITON (Organizar)", last_col=last_col)

    header_row = _create_info_section(ws, "S2", "Organizar — Un lugar para cada cosa y cada cosa en su lugar", last_col)

    # Data validations
    dv_frecuencia = DataValidation(type="list", formula1='"Muy frecuente,Frecuente,Ocasional,Raro"', allow_blank=True)
    dv_frecuencia.prompt = "Con qué frecuencia se usa"
    dv_frecuencia.promptTitle = "Frecuencia"
    dv_frecuencia.sqref = f"G{header_row + 1}:G{header_row + 15}"

    dv_metodo = DataValidation(type="list", formula1='"Etiqueta,Código color,Señal visual,Sombra/Contorno,Código numérico,Otro"', allow_blank=True)
    dv_metodo.prompt = "Método de identificación visual"
    dv_metodo.promptTitle = "Método"
    dv_metodo.sqref = f"I{header_row + 1}:I{header_row + 15}"

    dv_cercania = DataValidation(type="list", formula1='"Muy cerca (brazo),Cerca (1-3 pasos),Media distancia,Poco accesible"', allow_blank=True)
    dv_cercania.prompt = "Distancia al puesto de trabajo"
    dv_cercania.promptTitle = "Cercanía"
    dv_cercania.sqref = f"J{header_row + 1}:J{header_row + 15}"

    _create_inventory_table(ws, header_row, headers, col_widths,
                           [dv_frecuencia, dv_metodo, dv_cercania])

    ws.freeze_panes = f"D{header_row + 1}"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

    wb.properties.creator = "Z.ai"
    path = os.path.join(OUTPUT_DIR, "S2_Inventario_Necesarios_Seiton.xlsx")
    wb.save(path)
    print(f"✅ Generado: {path}")
    return path


# ============================================================
# S3 — Inventario de Puntos de Suciedad (Seiso - Limpiar)
# ============================================================
def create_s3_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "S3 Suciedad"

    headers = [
        "Nº", "Punto de Suciedad", "Ubicación", "Zona", "Tipo de Suciedad",
        "Nivel", "Fuente de Suciedad", "Método de Limpieza",
        "Frecuencia Limpieza", "Responsable", "Fecha Limpieza", "Observaciones"
    ]
    col_widths = [5, 22, 16, 12, 16, 12, 18, 18, 16, 14, 14, 24]

    last_col = len(headers) + 1
    setup_sheet(ws, title="Inventario de Puntos de Suciedad — SEISO (Limpiar)", last_col=last_col)

    header_row = _create_info_section(ws, "S3", "Limpiar — Identificar y eliminar fuentes de suciedad", last_col)

    # Data validations
    dv_tipo = DataValidation(type="list", formula1='"Polvo,Grasa,Mancha,Residuos,Humedad,Oxidación,Otro"', allow_blank=True)
    dv_tipo.prompt = "Tipo de suciedad encontrada"
    dv_tipo.promptTitle = "Tipo"
    dv_tipo.sqref = f"F{header_row + 1}:F{header_row + 15}"

    dv_nivel = DataValidation(type="list", formula1='"Leve,Moderado,Grave"', allow_blank=True)
    dv_nivel.prompt = "Nivel de suciedad"
    dv_nivel.promptTitle = "Nivel"
    dv_nivel.sqref = f"G{header_row + 1}:G{header_row + 15}"

    dv_fuente = DataValidation(type="list", formula1='"Proceso productivo,Medio ambiente,Falta de limpieza,Escape/Fuga,Desgaste,Derrame,Otro"', allow_blank=True)
    dv_fuente.prompt = "Origen de la suciedad"
    dv_fuente.promptTitle = "Fuente"
    dv_fuente.sqref = f"H{header_row + 1}:H{header_row + 15}"

    dv_metodo = DataValidation(type="list", formula1='"Aspirado,Fregado,Pulido,Desinfección,Aspiración,Reparación,Otro"', allow_blank=True)
    dv_metodo.prompt = "Método de limpieza necesario"
    dv_metodo.promptTitle = "Método"
    dv_metodo.sqref = f"I{header_row + 1}:I{header_row + 15}"

    dv_frecuencia = DataValidation(type="list", formula1='"Diaria,Tres veces/semana,Semanal,Quincenal,Mensual"', allow_blank=True)
    dv_frecuencia.prompt = "Frecuencia de limpieza requerida"
    dv_frecuencia.promptTitle = "Frecuencia"
    dv_frecuencia.sqref = f"J{header_row + 1}:J{header_row + 15}"

    _create_inventory_table(ws, header_row, headers, col_widths,
                           [dv_tipo, dv_nivel, dv_fuente, dv_metodo, dv_frecuencia])

    ws.freeze_panes = f"D{header_row + 1}"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

    wb.properties.creator = "Z.ai"
    path = os.path.join(OUTPUT_DIR, "S3_Inventario_Puntos_Suciedad_Seiso.xlsx")
    wb.save(path)
    print(f"✅ Generado: {path}")
    return path


# ============================================================
# S4 — Inventario de Estándares Implantados (Seiketsu - Estandarizar)
# ============================================================
def create_s4_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "S4 Estándares"

    headers = [
        "Nº", "Estándar / Procedimiento", "Ámbito / Proceso", "Zona",
        "Tipo de Estándar", "Estado", "Documentado",
        "Responsable Mant.", "Fecha Implantación", "Fecha Revisión",
        "Cumplimiento %", "Observaciones"
    ]
    col_widths = [5, 24, 16, 12, 16, 14, 12, 16, 14, 14, 14, 24]

    last_col = len(headers) + 1
    setup_sheet(ws, title="Inventario de Estándares Implantados — SEIKETSU (Estandarizar)", last_col=last_col)

    header_row = _create_info_section(ws, "S4", "Estandarizar — Crear normas y procedimientos para mantener las mejoras", last_col)

    # Data validations
    dv_tipo = DataValidation(type="list", formula1='"Visual,Procedimiento,Checklist,Señalización,Diagrama flujo,Registro,Otro"', allow_blank=True)
    dv_tipo.prompt = "Tipo de estándar implantado"
    dv_tipo.promptTitle = "Tipo"
    dv_tipo.sqref = f"F{header_row + 1}:F{header_row + 15}"

    dv_estado = DataValidation(type="list", formula1='"Implantado,En proceso,Pendiente"', allow_blank=True)
    dv_estado.prompt = "Estado del estándar"
    dv_estado.promptTitle = "Estado"
    dv_estado.sqref = f"G{header_row + 1}:G{header_row + 15}"

    dv_doc = DataValidation(type="list", formula1='"Sí,No,Parcialmente"', allow_blank=True)
    dv_doc.prompt = "¿Está documentado el estándar?"
    dv_doc.promptTitle = "Documentado"
    dv_doc.sqref = f"H{header_row + 1}:H{header_row + 15}"

    _create_inventory_table(ws, header_row, headers, col_widths,
                           [dv_tipo, dv_estado, dv_doc])

    ws.freeze_panes = f"D{header_row + 1}"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

    wb.properties.creator = "Z.ai"
    path = os.path.join(OUTPUT_DIR, "S4_Inventario_Estandares_Seiketsu.xlsx")
    wb.save(path)
    print(f"✅ Generado: {path}")
    return path


# ============================================================
# MAIN — Generate all 4 templates
# ============================================================
if __name__ == "__main__":
    print("=" * 60)
    print("Generando plantillas 5S — Inventarios")
    print("=" * 60)
    p1 = create_s1_template()
    p2 = create_s2_template()
    p3 = create_s3_template()
    p4 = create_s4_template()
    print("\n" + "=" * 60)
    print("✅ Todas las plantillas generadas correctamente:")
    print(f"   📋 S1: {p1}")
    print(f"   📋 S2: {p2}")
    print(f"   📋 S3: {p3}")
    print(f"   📋 S4: {p4}")
    print("=" * 60)
